# -*- coding: utf-8 -*-
"""
PhoneBook Telegram Bot (Arabic UI) + Admin Analytics (Private)
- Search + inline buttons
- SQLite stats (users + events)
- Admin panel:
  * Total users (all-time) + list of users who interacted with bot
  * Top 10 departments (all-time)
  * Top 15 users by usage (all-time)
  * Recent 25 active users (all-time) with time
  * Export submenu (CSV/XLSX) for: Summary, UsersAll, UsersUsed, TopDepts, TopUsers, FullPack
  * Broadcast message to all known users (confirmation)
Notes:
- "All members of a Telegram group" cannot be listed by bots via API. We can only list users who messaged/used the bot (tracked in DB).
"""

import os
import re
import io
import csv
import math
import sqlite3
import asyncio
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from telegram import (
    Update,
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler,
    ContextTypes, filters
)
from telegram.error import RetryAfter

# -------------------- Logging --------------------
logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# -------------------- Timezone (Karbala/Iraq) --------------------
try:
    from zoneinfo import ZoneInfo
    IRAQ_TZ = ZoneInfo("Asia/Baghdad")   # Karbala uses same timezone as Baghdad
except Exception:
    IRAQ_TZ = None

def now_iraq() -> datetime:
    if IRAQ_TZ:
        return datetime.now(IRAQ_TZ)
    return datetime.utcnow() + timedelta(hours=3)

def iso(dt: datetime) -> str:
    return dt.replace(microsecond=0).isoformat()

def fmt_ts(ts: str) -> str:
    if not ts:
        return "—"
    try:
        dt = datetime.fromisoformat(ts)
    except Exception:
        return ts
    try:
        if IRAQ_TZ:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=IRAQ_TZ)
            else:
                dt = dt.astimezone(IRAQ_TZ)
    except Exception:
        pass
    return dt.strftime("%Y-%m-%d  %H:%M:%S") + "  (Karbala)"

# -------------------- Paths --------------------
BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.getenv("DATA_DIR", BASE)
DB_PATH = os.path.join(BASE, "stats.db")

# -------------------- Admin --------------------
ADMIN_ID = 8099482759
ADMIN_USERNAME = "@ya_se91"

# -------------------- Branding --------------------
SIGNATURE = "\n────────────\nSource: CCTV – Yaseen Al-Tamimi"

INTRO_TEXT = (
    "👋 أهلاً بك في بوت أرقام مستشفى الإمام الحسن المجتبى (ع).\n\n"
    "📌 طريقة الاستخدام:\n"
    "• 📞 أرقام المستشفى: تصفّح الأقسام كمربعات.\n"
    "• 🔍 بحث بالاسم: اكتب أي جزء من اسم القسم.\n"
    "• ℹ️ عن البوت: معلومات عن البوت.\n\n"
    "✅ ملاحظة: الاقتراحات/التعديلات يرجى إرسالها إلى: " + ADMIN_USERNAME + "\n"
    + SIGNATURE
)

ABOUT_TEXT = (
    "ℹ️ عن البوت\n"
    "هذا البوت مخصص لعرض أرقام أقسام المستشفى بسرعة عبر البحث أو الأزرار.\n\n"
    "📩 لمزيد من الاستفسارات أو مقترحات التعديل:\n"
    f"{ADMIN_USERNAME}\n"
    + SIGNATURE
)

BROADCAST_TEXT = (
    "🌟 تحية طيبة من فريق مستشفى الإمام الحسن المجتبى (ع)\n\n"
    "نود معرفة رأيكم لتحسين بوت الأرقام:\n"
    "هل لديكم أي اقتراحات أو تعديلات تحبون نضيفها؟\n\n"
    f"📩 أرسلوا اقتراحاتكم إلى: {ADMIN_USERNAME}\n"
    + SIGNATURE
)

# -------------------- UI Keyboards --------------------
MAIN_KB = ReplyKeyboardMarkup(
    [
        [KeyboardButton("📞 أرقام المستشفى")],
        [KeyboardButton("🔍 بحث بالاسم")],
        [KeyboardButton("ℹ️ عن البوت")],
        [KeyboardButton("◀️ رجوع للقائمة")]
    ],
    resize_keyboard=True
)

GRID_COLS = 3
PAGE_SIZE_ALL = 24
PAGE_SIZE_SRCH = 21

# -------------------- Arabic normalize --------------------
ARABIC_DIAC = re.compile(r"[ًٌٍَُِّْـ]")

def strip_diacritics(s: str) -> str:
    return ARABIC_DIAC.sub("", s or "")

def normalize_arabic(s: str) -> str:
    s = str(s or "")
    s = s.replace("\u200f","").replace("\u200e","").replace("\ufeff","").strip()
    s = strip_diacritics(s)
    s = s.replace("آ","ا").replace("أ","ا").replace("إ","ا")
    s = s.replace("ى","ي").replace("ة","ه")
    s = re.sub(r"[^\w\s\u0600-\u06FF]"," ", s)
    s = re.sub(r"\s+"," ", s).strip()
    return s.upper()

# -------------------- Excel loading --------------------
DEPT_CANDIDATES  = ["القسم","قسم","الاسم","اسم القسم"]
PHONE_CANDIDATES = ["رقم الهاتف","الهاتف","رقم","موبايل","Phone"]

display_rows: List[Tuple[str, str]] = []
departments: List[str] = []
phonebook: Dict[str, str] = {}

def list_excel_files(folder: str) -> List[str]:
    try:
        return [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    except Exception:
        return []

def read_headers(ws) -> List[str]:
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c or "").strip() for c in row]
    return []

def find_col_idx(headers: List[str], candidates: List[str]) -> Optional[int]:
    H = [normalize_arabic(h) for h in headers]
    C = [normalize_arabic(c) for c in candidates]
    for i, h in enumerate(H):
        if h in C:
            return i
    for i, h in enumerate(H):
        for c in C:
            if c in h:
                return i
    return None

def load_phonebook() -> Tuple[int, str]:
    global display_rows, departments, phonebook
    display_rows, departments, phonebook = [], [], {}
    files = list_excel_files(DATA_DIR)
    if not files:
        return 0, f"❌ ماكو ملفات ‎.xlsx داخل: {DATA_DIR}"
    total = 0
    for path in files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = read_headers(ws)
            if not headers:
                wb.close()
                continue
            di = find_col_idx(headers, DEPT_CANDIDATES)
            pi = find_col_idx(headers, PHONE_CANDIDATES)
            if di is None or pi is None:
                wb.close()
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                dept = str(row[di] if di < len(row) and row[di] is not None else "").strip()
                phone = str(row[pi] if pi < len(row) and row[pi] is not None else "").strip()
                if not dept:
                    continue
                display_rows.append((dept, phone))
                phonebook[normalize_arabic(dept)] = phone
                total += 1
            wb.close()
        except Exception as e:
            logging.exception(f"Excel load error in {path}: {e}")

    display_rows.sort(key=lambda x: x[0])
    departments = [d for d, _ in display_rows]
    return total, (f"✅ تم تحميل {total} سجل." if total else "❌ لم يتم تحميل أي سجل.")

# -------------------- DB --------------------
def db_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn

def init_db():
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            first_seen TEXT NOT NULL,
            last_seen  TEXT NOT NULL,
            username   TEXT,
            full_name  TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            chat_id INTEGER,
            event_type TEXT NOT NULL,
            dept TEXT,
            query TEXT,
            extra TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_ts ON events(ts)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_type ON events(event_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_dept ON events(dept)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_user ON events(user_id)")
    conn.commit()
    conn.close()

def upsert_user(user) -> None:
    if not user:
        return
    uid = user.id
    username = user.username or ""
    full_name = (user.full_name or "").strip()
    t = iso(now_iraq())
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT user_id FROM users WHERE user_id=?", (uid,))
    row = cur.fetchone()
    if row:
        cur.execute(
            "UPDATE users SET last_seen=?, username=?, full_name=? WHERE user_id=?",
            (t, username, full_name, uid)
        )
    else:
        cur.execute(
            "INSERT INTO users(user_id, first_seen, last_seen, username, full_name) VALUES(?,?,?,?,?)",
            (uid, t, t, username, full_name)
        )
    conn.commit()
    conn.close()

def log_event(event_type: str, user_id: int, chat_id: Optional[int], dept: str = "", query: str = "", extra: str = "") -> None:
    t = iso(now_iraq())
    conn = db_conn()
    conn.execute(
        "INSERT INTO events(ts, user_id, chat_id, event_type, dept, query, extra) VALUES(?,?,?,?,?,?,?)",
        (t, user_id, chat_id if chat_id is not None else None, event_type, dept or "", query or "", extra or "")
    )
    conn.commit()
    conn.close()

def is_admin(update: Update) -> bool:
    u = update.effective_user
    return bool(u and u.id == ADMIN_ID)

# -------------------- Helpers: send safe --------------------
async def safe_send_text(msg, text: str, reply_markup=None):
    text = f"{text}{SIGNATURE}"
    try:
        return await msg.reply_text(text, reply_markup=reply_markup)
    except RetryAfter as e:
        await asyncio.sleep(e.retry_after + 1)
        return await msg.reply_text(text, reply_markup=reply_markup)

async def safe_send_doc(msg, file_bytes: bytes, filename: str, caption: str):
    try:
        bio = io.BytesIO(file_bytes)
        bio.name = filename
        return await msg.reply_document(document=bio, filename=filename, caption=caption)
    except RetryAfter as e:
        await asyncio.sleep(e.retry_after + 1)
        bio = io.BytesIO(file_bytes)
        bio.name = filename
        return await msg.reply_document(document=bio, filename=filename, caption=caption)

# -------------------- Search / Grids --------------------
def search_indices(query: str) -> List[int]:
    qn = normalize_arabic(query)
    if not qn:
        return []
    matches = []
    for i, name in enumerate(departments):
        if qn in normalize_arabic(name):
            matches.append(i)
    return matches

def build_grid(indices: List[int], page: int, page_size: int, cols: int, mode: str) -> InlineKeyboardMarkup:
    total = len(indices)
    pages = max(1, math.ceil(total / page_size))
    page  = max(0, min(page, pages - 1))
    start, end = page * page_size, min(page * page_size + page_size, total)
    slice_idx = indices[start:end]

    rows, row = [], []
    for idx in slice_idx:
        name = departments[idx]
        row.append(InlineKeyboardButton(name, callback_data=f"dept:{idx}"))
        if len(row) == cols:
            rows.append(row)
            row = []
    if row:
        rows.append(row)

    if pages > 1:
        ctrl = []
        if page > 0:
            ctrl.append(InlineKeyboardButton("⬅️ السابق", callback_data=f"{mode}:{page-1}"))
        ctrl.append(InlineKeyboardButton(f"صفحة {page+1}/{pages}", callback_data="noop"))
        if page < pages-1:
            ctrl.append(InlineKeyboardButton("التالي ➡️", callback_data=f"{mode}:{page+1}"))
        rows.append(ctrl)

    rows.append([InlineKeyboardButton("◀️ رجوع للقائمة", callback_data="home")])
    return InlineKeyboardMarkup(rows)

def grid_all(page: int = 0) -> InlineKeyboardMarkup:
    return build_grid(list(range(len(departments))), page, PAGE_SIZE_ALL, GRID_COLS, "allp")

def grid_search(matches: List[int], page: int = 0) -> InlineKeyboardMarkup:
    return build_grid(matches, page, PAGE_SIZE_SRCH, GRID_COLS, "srchp")

# -------------------- Admin Panel UI --------------------
def admin_menu() -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("📊 ملخص شامل (من البداية)", callback_data="adm:summary")],
        [InlineKeyboardButton("🏆 Top 10 أقسام (من البداية)", callback_data="adm:top_depts")],
        [InlineKeyboardButton("👥 عدد المستخدمين + قائمة المستخدمين", callback_data="adm:users_list:0")],
        [InlineKeyboardButton("👥 Top 15 مستخدم استخداماً", callback_data="adm:top_users")],
        [InlineKeyboardButton("🕒 آخر 25 مستخدم نشط", callback_data="adm:recent25")],
        [InlineKeyboardButton("📥 تصدير التقارير", callback_data="adm:export_menu")],
        [InlineKeyboardButton("📣 إرسال رسالة ترحيب/اقتراحات للجميع", callback_data="adm:broadcast_confirm")],
        [InlineKeyboardButton("◀️ رجوع للقائمة", callback_data="home")],
    ]
    return InlineKeyboardMarkup(rows)

def export_menu() -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("📄 Summary (CSV)", callback_data="adm:export:summary:csv"),
         InlineKeyboardButton("📄 Summary (XLSX)", callback_data="adm:export:summary:xlsx")],
        [InlineKeyboardButton("👥 Users All (CSV)", callback_data="adm:export:users_all:csv"),
         InlineKeyboardButton("👥 Users All (XLSX)", callback_data="adm:export:users_all:xlsx")],
        [InlineKeyboardButton("✅ Users Used (CSV)", callback_data="adm:export:users_used:csv"),
         InlineKeyboardButton("✅ Users Used (XLSX)", callback_data="adm:export:users_used:xlsx")],
        [InlineKeyboardButton("🏆 Top Depts (CSV)", callback_data="adm:export:top_depts:csv"),
         InlineKeyboardButton("🏆 Top Depts (XLSX)", callback_data="adm:export:top_depts:xlsx")],
        [InlineKeyboardButton("👥 Top Users (CSV)", callback_data="adm:export:top_users:csv"),
         InlineKeyboardButton("👥 Top Users (XLSX)", callback_data="adm:export:top_users:xlsx")],
        [InlineKeyboardButton("📦 Full Pack (CSV)", callback_data="adm:export:full:csv"),
         InlineKeyboardButton("📦 Full Pack (XLSX)", callback_data="adm:export:full:xlsx")],
        [InlineKeyboardButton("⬅️ رجوع", callback_data="adm:back_admin")],
    ]
    return InlineKeyboardMarkup(rows)

# -------------------- Admin queries --------------------
def q_total_users() -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    n = cur.fetchone()[0] or 0
    conn.close()
    return n

def q_last_activity_ts() -> str:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT MAX(ts) FROM events")
    ts = cur.fetchone()[0] or ""
    conn.close()
    return ts

def q_top10_depts() -> List[Tuple[str, int]]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT dept, COUNT(*) AS c
        FROM events
        WHERE event_type IN ('dept_select','search_hit') AND dept <> ''
        GROUP BY dept
        ORDER BY c DESC
        LIMIT 10
    """)
    rows = cur.fetchall()
    conn.close()
    return [(r[0], int(r[1])) for r in rows]

def q_top15_users() -> List[Tuple[int, int, str, str, str, str]]:
    """
    returns (user_id, usage_count, full_name, username, first_used, last_used)
    """
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, COUNT(*) AS c, MIN(ts) AS first_used, MAX(ts) AS last_used
        FROM events
        WHERE event_type IN ('dept_select','search_hit','search_text')
        GROUP BY user_id
        ORDER BY c DESC
        LIMIT 15
    """)
    rows = cur.fetchall()

    out = []
    for uid, c, first_used, last_used in rows:
        cur.execute("SELECT full_name, username FROM users WHERE user_id=?", (uid,))
        urow = cur.fetchone()
        full_name = (urow[0] if urow and urow[0] else "").strip()
        username = (urow[1] if urow and urow[1] else "").strip()
        out.append((int(uid), int(c), full_name, username, first_used or "", last_used or ""))
    conn.close()
    return out

def q_recent25_active() -> List[Tuple[int, str, str, str]]:
    """
    last 25 users by last event timestamp
    returns (user_id, full_name, username, last_used)
    """
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, MAX(ts) AS last_used
        FROM events
        GROUP BY user_id
        ORDER BY last_used DESC
        LIMIT 25
    """)
    rows = cur.fetchall()

    out = []
    for uid, last_used in rows:
        cur.execute("SELECT full_name, username FROM users WHERE user_id=?", (uid,))
        urow = cur.fetchone()
        full_name = (urow[0] if urow and urow[0] else "").strip()
        username = (urow[1] if urow and urow[1] else "").strip()
        out.append((int(uid), full_name, username, last_used or ""))
    conn.close()
    return out

def q_users_page(offset: int, limit: int = 50) -> List[Tuple[int, str, str, str, str]]:
    """
    returns list of (user_id, full_name, username, first_seen, last_seen) ordered by first_seen desc
    """
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, full_name, username, first_seen, last_seen
        FROM users
        ORDER BY first_seen DESC
        LIMIT ? OFFSET ?
    """, (limit, offset))
    rows = cur.fetchall()
    conn.close()
    return [(int(uid), (fn or ""), (un or ""), (fs or ""), (ls or "")) for uid, fn, un, fs, ls in rows]

def q_users_used_all() -> List[Tuple[int, str, str, str, str, int]]:
    """
    Users who used bot (has any event). returns:
    (user_id, full_name, username, first_used, last_used, usage_count)
    """
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, COUNT(*) AS c, MIN(ts) AS first_used, MAX(ts) AS last_used
        FROM events
        GROUP BY user_id
        ORDER BY first_used ASC
    """)
    rows = cur.fetchall()

    out = []
    for uid, c, first_used, last_used in rows:
        cur.execute("SELECT full_name, username FROM users WHERE user_id=?", (uid,))
        urow = cur.fetchone()
        full_name = (urow[0] if urow and urow[0] else "").strip()
        username = (urow[1] if urow and urow[1] else "").strip()
        out.append((int(uid), full_name, username, first_used or "", last_used or "", int(c)))
    conn.close()
    return out

# -------------------- Export builders (CSV/XLSX) --------------------
def build_summary_rows() -> List[Tuple[str, str]]:
    total_users = q_total_users()
    last_act = fmt_ts(q_last_activity_ts())
    return [
        ("Bot", "Imam Al-Hasan Al-Mujtaba Hospital PhoneBook"),
        ("GeneratedAt", fmt_ts(iso(now_iraq()))),
        ("TotalUsers", str(total_users)),
        ("LastActivity", last_act),
    ]

def build_top_depts_rows() -> List[Tuple[int, str, int]]:
    rows = q_top10_depts()
    out = []
    for i, (dept, c) in enumerate(rows, 1):
        out.append((i, dept, c))
    return out

def build_top_users_rows() -> List[Tuple[int, int, str, str, int, str, str]]:
    rows = q_top15_users()
    out = []
    for i, (uid, c, full_name, username, first_used, last_used) in enumerate(rows, 1):
        out.append((i, uid, full_name, ("@" + username) if username else "", c, fmt_ts(first_used), fmt_ts(last_used)))
    return out

def build_users_all_rows() -> List[Tuple[int, str, str, str, str]]:
    # all users from users table (ordered by first_seen desc)
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, full_name, username, first_seen, last_seen
        FROM users
        ORDER BY first_seen ASC
    """)
    rows = cur.fetchall()
    conn.close()
    out = []
    for uid, fn, un, fs, ls in rows:
        out.append((int(uid), (fn or ""), ("@" + un) if un else "", fmt_ts(fs or ""), fmt_ts(ls or "")))
    return out

def build_users_used_rows() -> List[Tuple[int, str, str, str, str, int]]:
    rows = q_users_used_all()
    out = []
    for uid, fn, un, first_used, last_used, c in rows:
        out.append((uid, fn, ("@" + un) if un else "", fmt_ts(first_used), fmt_ts(last_used), c))
    return out

def to_csv_bytes(sheet_name: str, headers: List[str], rows: List[Tuple]) -> bytes:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow([sheet_name])
    w.writerow([])
    w.writerow(headers)
    for r in rows:
        w.writerow(list(r))
    return out.getvalue().encode("utf-8-sig")

def xlsx_bytes(sheets: List[Tuple[str, List[str], List[Tuple]]]) -> bytes:
    wb = Workbook()
    # remove default
    wb.remove(wb.active)

    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for r in rows:
            ws.append(list(r))

        # simple formatting: bold header + freeze
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # autosize columns (approx)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(45, max(10, max_len + 2))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def build_export(kind: str, fmt: str) -> Tuple[str, bytes]:
    """
    kind: summary | users_all | users_used | top_depts | top_users | full
    fmt: csv | xlsx
    """
    summary = build_summary_rows()
    top_depts = build_top_depts_rows()
    top_users = build_top_users_rows()
    users_all = build_users_all_rows()
    users_used = build_users_used_rows()

    if kind == "summary":
        if fmt == "csv":
            filename = "summary.csv"
            data = to_csv_bytes("Summary", ["Key", "Value"], summary)
            return filename, data
        else:
            filename = "summary.xlsx"
            data = xlsx_bytes([("Summary", ["Key", "Value"], summary)])
            return filename, data

    if kind == "top_depts":
        if fmt == "csv":
            filename = "top10_departments.csv"
            data = to_csv_bytes("Top10Departments", ["Rank", "Department", "SearchCount"], top_depts)
            return filename, data
        else:
            filename = "top10_departments.xlsx"
            data = xlsx_bytes([("Top10Departments", ["Rank", "Department", "SearchCount"], top_depts)])
            return filename, data

    if kind == "top_users":
        headers = ["Rank", "UserID", "Name", "Username", "UsageCount", "FirstUsed", "LastUsed"]
        if fmt == "csv":
            filename = "top15_users.csv"
            data = to_csv_bytes("Top15Users", headers, top_users)
            return filename, data
        else:
            filename = "top15_users.xlsx"
            data = xlsx_bytes([("Top15Users", headers, top_users)])
            return filename, data

    if kind == "users_all":
        headers = ["UserID", "Name", "Username", "FirstSeen", "LastSeen"]
        if fmt == "csv":
            filename = "users_all.csv"
            data = to_csv_bytes("UsersAll", headers, users_all)
            return filename, data
        else:
            filename = "users_all.xlsx"
            data = xlsx_bytes([("UsersAll", headers, users_all)])
            return filename, data

    if kind == "users_used":
        headers = ["UserID", "Name", "Username", "FirstUsed", "LastUsed", "UsageCount"]
        if fmt == "csv":
            filename = "users_used.csv"
            data = to_csv_bytes("UsersUsed", headers, users_used)
            return filename, data
        else:
            filename = "users_used.xlsx"
            data = xlsx_bytes([("UsersUsed", headers, users_used)])
            return filename, data

    # full pack
    if fmt == "csv":
        # one csv file with sections
        out = io.StringIO()
        w = csv.writer(out)

        w.writerow(["Summary"]); w.writerow([])
        w.writerow(["Key","Value"])
        for r in summary: w.writerow(list(r))
        w.writerow([]); w.writerow([])

        w.writerow(["Top10Departments"]); w.writerow([])
        w.writerow(["Rank","Department","SearchCount"])
        for r in top_depts: w.writerow(list(r))
        w.writerow([]); w.writerow([])

        w.writerow(["Top15Users"]); w.writerow([])
        w.writerow(["Rank","UserID","Name","Username","UsageCount","FirstUsed","LastUsed"])
        for r in top_users: w.writerow(list(r))
        w.writerow([]); w.writerow([])

        w.writerow(["UsersAll"]); w.writerow([])
        w.writerow(["UserID","Name","Username","FirstSeen","LastSeen"])
        for r in users_all: w.writerow(list(r))
        w.writerow([]); w.writerow([])

        w.writerow(["UsersUsed"]); w.writerow([])
        w.writerow(["UserID","Name","Username","FirstUsed","LastUsed","UsageCount"])
        for r in users_used: w.writerow(list(r))

        return "full_report.csv", out.getvalue().encode("utf-8-sig")

    else:
        sheets = [
            ("Summary", ["Key","Value"], summary),
            ("Top10Departments", ["Rank","Department","SearchCount"], top_depts),
            ("Top15Users", ["Rank","UserID","Name","Username","UsageCount","FirstUsed","LastUsed"], top_users),
            ("UsersAll", ["UserID","Name","Username","FirstSeen","LastSeen"], users_all),
            ("UsersUsed", ["UserID","Name","Username","FirstUsed","LastUsed","UsageCount"], users_used),
        ]
        return "full_report.xlsx", xlsx_bytes(sheets)

# -------------------- Handlers --------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("start", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    await update.message.reply_text(INTRO_TEXT, reply_markup=MAIN_KB)

async def about_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("about", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    await safe_send_text(update.message, ABOUT_TEXT, reply_markup=MAIN_KB)

async def reload_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("reload", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    n, msg = load_phonebook()
    await safe_send_text(update.message, msg, reply_markup=MAIN_KB)

async def admin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("admin_open", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    if not is_admin(update):
        await safe_send_text(update.message, "⛔️ غير مصرح.", reply_markup=MAIN_KB)
        return
    await safe_send_text(update.message, "👑 لوحة الإدارة والإحصائيات:", reply_markup=admin_menu())

async def list_depts(update: Update, page: int = 0):
    if not departments:
        await safe_send_text(update.message, "❌ لا توجد سجلات. استخدم /reload بعد التأكد من ملف الإكسل.", reply_markup=MAIN_KB)
        return
    await update.message.reply_text("اختر القسم من القائمة:", reply_markup=grid_all(page))

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    uid = update.effective_user.id
    chat_id = update.effective_chat.id if update.effective_chat else None
    txt = (update.message.text or "").strip()

    if txt == "📞 أرقام المستشفى":
        log_event("open_list", uid, chat_id)
        await list_depts(update, 0)
        return

    if txt == "🔍 بحث بالاسم":
        log_event("prompt_search", uid, chat_id)
        await safe_send_text(update.message, "✍️ اكتب أي جزء من اسم القسم.", reply_markup=MAIN_KB)
        return

    if txt == "ℹ️ عن البوت":
        log_event("about_btn", uid, chat_id)
        await safe_send_text(update.message, ABOUT_TEXT, reply_markup=MAIN_KB)
        return

    if txt == "◀️ رجوع للقائمة":
        log_event("back_home", uid, chat_id)
        await update.message.reply_text(INTRO_TEXT, reply_markup=MAIN_KB)
        return

    # text search
    matches = search_indices(txt)
    log_event("search_text", uid, chat_id, query=txt, extra=f"matches={len(matches)}")

    if not matches:
        log_event("not_found", uid, chat_id, query=txt)
        await safe_send_text(update.message, "❌ لم يتم العثور على هذا القسم.", reply_markup=MAIN_KB)
        return

    if len(matches) == 1:
        idx = matches[0]
        name = departments[idx]
        num = phonebook.get(normalize_arabic(name), "")
        log_event("search_hit", uid, chat_id, dept=name, query=txt)
        await safe_send_text(update.message, f"✅ {name} — {num if num else '—'}", reply_markup=MAIN_KB)
        return

    context.user_data["last_search_indices"] = matches
    await update.message.reply_text("🔎 تم العثور على عدة نتائج، اختر القسم:", reply_markup=grid_search(matches, 0))

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    data = q.data if q else ""
    uid = update.effective_user.id if update.effective_user else None
    chat_id = update.effective_chat.id if update.effective_chat else None

    await q.answer()

    # HOME
    if data == "home":
        try:
            await q.message.edit_text(INTRO_TEXT)
        except Exception:
            pass
        await q.message.reply_text("رجعت للقائمة الرئيسية.", reply_markup=MAIN_KB)
        return

    if data == "noop":
        return

    # pagination
    if data.startswith("allp:"):
        page = int(data.split(":")[1])
        await q.message.edit_text("اختر القسم من القائمة:", reply_markup=grid_all(page))
        return

    if data.startswith("srchp:"):
        page = int(data.split(":")[1])
        matches = context.user_data.get("last_search_indices", [])
        await q.message.edit_text("🔎 تم العثور على عدة نتائج، اختر القسم:", reply_markup=grid_search(matches, page))
        return

    # dept
    if data.startswith("dept:"):
        idx = int(data.split(":")[1])
        if 0 <= idx < len(departments):
            name = departments[idx]
            num = phonebook.get(normalize_arabic(name), "")
            upsert_user(update.effective_user)
            log_event("dept_select", uid, chat_id, dept=name)
            await q.message.reply_text(f"📞 {name} — {num if num else '—'}{SIGNATURE}")
        else:
            await q.message.reply_text(f"خيار غير صالح.{SIGNATURE}")
        return

    # ADMIN callbacks
    if data.startswith("adm:"):
        if uid != ADMIN_ID:
            await q.message.reply_text(f"⛔️ غير مصرح.{SIGNATURE}", reply_markup=MAIN_KB)
            return

        # back to admin
        if data == "adm:back_admin":
            await q.message.reply_text("👑 لوحة الإدارة والإحصائيات:", reply_markup=admin_menu())
            return

        if data == "adm:summary":
            total_users = q_total_users()
            last_act = fmt_ts(q_last_activity_ts())
            top_depts = q_top10_depts()
            top_users = q_top15_users()

            lines = [
                "📊 ملخص شامل (من البداية)",
                f"• 👥 عدد المستخدمين الكلي: {total_users}",
                f"• 🕒 آخر نشاط: {last_act}",
                "",
                "🏆 Top 10 أقسام:",
            ]
            if top_depts:
                for i, (d, c) in enumerate(top_depts, 1):
                    lines.append(f"{i}) {d} — {c}")
            else:
                lines.append("— لا توجد بيانات كافية بعد —")

            lines.append("")
            lines.append("👥 Top 15 مستخدم استخداماً:")
            if top_users:
                for i, (u_id, c, full_name, username, first_used, last_used) in enumerate(top_users, 1):
                    label = full_name if full_name else (("@" + username) if username else str(u_id))
                    lines.append(f"{i}) {label} — {c} | آخر: {fmt_ts(last_used)}")
            else:
                lines.append("— لا توجد بيانات كافية بعد —")

            await q.message.reply_text("\n".join(lines) + SIGNATURE, reply_markup=admin_menu())
            return

        if data == "adm:top_depts":
            rows = q_top10_depts()
            lines = ["🏆 Top 10 أقسام (من البداية)"]
            if not rows:
                lines.append("— لا توجد بيانات كافية بعد —")
            else:
                for i, (dept, c) in enumerate(rows, 1):
                    lines.append(f"{i}) {dept} — {c}")
            await q.message.reply_text("\n".join(lines) + SIGNATURE, reply_markup=admin_menu())
            return

        if data.startswith("adm:users_list:"):
            page = int(data.split(":")[2])
            total = q_total_users()
            page_size = 50
            offset = page * page_size
            rows = q_users_page(offset, page_size)

            lines = [f"👥 عدد المستخدمين الكلي: {total}", ""]
            if not rows:
                lines.append("— لا توجد بيانات بعد —")
            else:
                for uid2, fn, un, fs, ls in rows:
                    name = fn if fn else str(uid2)
                    username = ("@" + un) if un else "—"
                    lines.append(f"• {name}  |  {username}  |  آخر: {fmt_ts(ls)}")

            # pagination buttons
            nav = []
            if page > 0:
                nav.append(InlineKeyboardButton("⬅️ السابق", callback_data=f"adm:users_list:{page-1}"))
            # if there might be next
            if len(rows) == page_size:
                nav.append(InlineKeyboardButton("التالي ➡️", callback_data=f"adm:users_list:{page+1}"))

            kb_rows = []
            if nav:
                kb_rows.append(nav)
            kb_rows.append([InlineKeyboardButton("📥 تصدير UsersAll", callback_data="adm:export:users_all:xlsx")])
            kb_rows.append([InlineKeyboardButton("⬅️ رجوع", callback_data="adm:back_admin")])

            await q.message.reply_text("\n".join(lines) + SIGNATURE, reply_markup=InlineKeyboardMarkup(kb_rows))
            return

        if data == "adm:top_users":
            rows = q_top15_users()
            lines = ["👥 Top 15 مستخدم استخداماً (من البداية)"]
            if not rows:
                lines.append("— لا توجد بيانات كافية بعد —")
            else:
                for i, (u_id, c, full_name, username, first_used, last_used) in enumerate(rows, 1):
                    name = full_name if full_name else str(u_id)
                    uname = ("@" + username) if username else "—"
                    lines.append(f"{i}) {name} | {uname}")
                    lines.append(f"    • استخدام: {c}  |  أول: {fmt_ts(first_used)}  |  آخر: {fmt_ts(last_used)}")
            await q.message.reply_text("\n".join(lines) + SIGNATURE, reply_markup=admin_menu())
            return

        if data == "adm:recent25":
            rows = q_recent25_active()
            lines = ["🕒 آخر 25 مستخدم نشط (من البداية)"]
            if not rows:
                lines.append("— لا توجد بيانات كافية بعد —")
            else:
                for i, (u_id, full_name, username, last_used) in enumerate(rows, 1):
                    name = full_name if full_name else str(u_id)
                    uname = ("@" + username) if username else "—"
                    lines.append(f"{i}) {name} | {uname} | آخر: {fmt_ts(last_used)}")
            await q.message.reply_text("\n".join(lines) + SIGNATURE, reply_markup=admin_menu())
            return

        if data == "adm:export_menu":
            await q.message.reply_text("📥 اختر نوع التصدير:", reply_markup=export_menu())
            return

        if data.startswith("adm:export:"):
            # adm:export:<kind>:<fmt>
            _, _, kind, fmt = data.split(":")
            try:
                filename, bytes_data = build_export(kind, fmt)
            except Exception as e:
                logging.exception(f"Export error {kind}/{fmt}: {e}")
                await q.message.reply_text(f"❌ فشل التصدير: {e}{SIGNATURE}", reply_markup=export_menu())
                return

            caption = f"📎 تقرير جاهز: {filename}\nGenerated: {fmt_ts(iso(now_iraq()))}"
            await safe_send_doc(q.message, bytes_data, filename, caption)
            return

        if data == "adm:broadcast_confirm":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ إرسال الآن", callback_data="adm:broadcast_send"),
                 InlineKeyboardButton("❌ إلغاء", callback_data="adm:back_admin")]
            ])
            await q.message.reply_text("⚠️ سيتم إرسال رسالة ترحيب/اقتراحات لجميع المستخدمين المسجلين.\nهل تريد المتابعة؟" + SIGNATURE, reply_markup=kb)
            return

        if data == "adm:broadcast_send":
            # send to all users in users table
            conn = db_conn()
            cur = conn.cursor()
            cur.execute("SELECT user_id FROM users")
            users = [int(r[0]) for r in cur.fetchall()]
            conn.close()

            ok = 0
            fail = 0
            for u_id in users:
                try:
                    await context.bot.send_message(chat_id=u_id, text=BROADCAST_TEXT)
                    ok += 1
                    await asyncio.sleep(0.05)
                except Exception:
                    fail += 1
                    await asyncio.sleep(0.05)

            await q.message.reply_text(f"✅ تم الإرسال.\nنجح: {ok}\nفشل: {fail}{SIGNATURE}", reply_markup=admin_menu())
            return

        # fallback
        await q.message.reply_text(f"خيار غير معروف.{SIGNATURE}", reply_markup=admin_menu())
        return

# -------------------- Token --------------------
def read_token() -> Optional[str]:
    tok = os.getenv("TELEGRAM_BOT_TOKEN")
    if tok:
        return tok.strip()
    path = os.path.join(BASE, "token.txt")
    if os.path.exists(path):
        return open(path, "r", encoding="utf-8").read().strip()
    return None

# -------------------- Main --------------------
if __name__ == "__main__":
    init_db()
    n, msg = load_phonebook()
    logging.info(msg)

    token = read_token()
    if not token:
        print("❌ لا يوجد توكن: ضع TELEGRAM_BOT_TOKEN أو token.txt.")
        raise SystemExit(1)

    app = ApplicationBuilder().token(token).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("about", about_cmd))
    app.add_handler(CommandHandler("reload", reload_cmd))
    app.add_handler(CommandHandler("admin", admin_cmd))

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(CallbackQueryHandler(on_callback))

    print("📞 PhoneBook Bot running…")
    app.run_polling()
